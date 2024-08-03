# -*- coding: utf-8 -*-
"""
Created on Thu Aug  1 10:50:32 2024

@author: M67B363
"""

from PyQt5.QtWidgets import QVBoxLayout, QMainWindow, QApplication, QLabel, QPushButton
from PyQt5.QtGui import QFont
import sys
from calculators import get_largest_year, pensioenjaar_calculator, get_input_vanaf_prognosejaar, creeer_pensioenjaar_parameters
from inputfiledialog import choose_file
from ontwikkel import ontwikkel_scenario
from parsers import parse_sheet
from parsers2 import read_and_parse_twod
from parsers3 import parse_fourds, vertical_parameters
import multiprocessing as mp
import logging
import numpy as np
import openpyxl
from datetime import datetime
from collections import defaultdict

def get_percentiles_now(year_data, percentiles):
    sorted_data = sorted(year_data.items(), key=lambda x: (x[1]['nominal_benefit'] if x[1]['nominal_benefit'] > 0 else x[1]['total_capital']))
    result = {}
    for p in percentiles:
        index = int(np.percentile(range(len(sorted_data)), p))
        scenario, data = sorted_data[index]
        result[f"{p}th"] = {'scenario': scenario, 'results': data}
    return result

def stap_1(nummer, eerste_berekeningsjaar, scenario_aantal, parameters, fourd, twod, endprognoseyear, pensioenleeftijd):
    try:
        # Log the current participant number and length of parameters
        logging.info(f"Processing participant {nummer}. Parameters list length: {len(parameters)}")
        
        # Check if index is in range
        if nummer < 0 or nummer >= len(parameters):
            logging.error(f"Participant index {nummer} out of range for parameters list of length {len(parameters)}")
            return None
        
        logging.info(f"Valid participant index {nummer}.")
        
        participant_parameters = parameters[nummer]
        logging.info(f"Parameters for participant {nummer}: {participant_parameters}")

        results = []
        for scenario in range(1, scenario_aantal + 1):
            logging.debug(f"Processing scenario {scenario} for participant {nummer}")
            pensioenjaar = pensioenjaar_calculator(participant_parameters, pensioenleeftijd)
            berekeningsjaar = participant_parameters['calculationdate'].year
            try:
                result = process_scenario(
                    eerste_berekeningsjaar,
                    scenario, 
                    nummer, 
                    participant_parameters, 
                    fourd, 
                    twod, 
                    endprognoseyear, 
                    pensioenleeftijd, 
                    pensioenjaar, 
                    berekeningsjaar
                )
                results.append(result)
            except Exception as e:
                logging.error(f"Error processing scenario {scenario} for participant {nummer}: {e}")
                raise e
        restructured_results = defaultdict(lambda: defaultdict(dict))

        for result in results:
            scenario = result["scenario"]
            for entry in result["scenarioresults"]:
                year = entry["jaar"]
                nominal_benefit = entry["nominal_benefit"]
                nominal_benefit_sr = entry["nominal_benefit_sr"]
                real_benefit = entry["real_benefit"]
                total_capital = entry["total_capital"]
                savings_op = entry["savings_op"]
                savings_hon = entry["savings_hon"]
                leeftijd = entry["leeftijd"]
                restructured_results[year][scenario] = {
                    "nominal_benefit": nominal_benefit,
                    "nominal_benefit_sr": nominal_benefit_sr,
                    "real_benefit": real_benefit,
                    "total_capital": total_capital,
                    "savings_op": savings_op,
                    "savings_hon": savings_hon,
                    "leeftijd": leeftijd
                    }
        
        final_result = {}
        percentiles = [5, 50, 95]

        for year, year_data in restructured_results.items():
            final_result[year] = get_percentiles_now(year_data, percentiles)

        
        return {
            "deelnemer": nummer + 1,
            "results": final_result
        }
    except IndexError as e:
        logging.error(f"IndexError for participant {nummer}: {e}")
        return None
    

def stap_2(nummer, eerste_berekeningsjaar, scenario_aantal, participant_parameters, fourd, twod, endprognoseyear, pensioenleeftijd, result):
    try:
        
        result = result[nummer]['results']
        #percentile_scenarios = get_percentile_scenarios(results)
        participant_parameters = participant_parameters[nummer]
        pensioenjaar = pensioenjaar_calculator(participant_parameters, pensioenleeftijd)
        berekeningsjaar = participant_parameters['calculationdate'].year
        if pensioenjaar < berekeningsjaar:
            pensioenjaar = berekeningsjaar
        if pensioenjaar >= berekeningsjaar:
            pensioenjaar = min(get_largest_year(result), pensioenjaar)
        
        mediaan_scenario = result[pensioenjaar]["50th"]
        porgnosejaar = pensioenjaar - participant_parameters['calculationdate'].year + 1
        twods, fourds = get_input_vanaf_prognosejaar(twod, fourd, porgnosejaar)
        pensioenjaar_parameters = creeer_pensioenjaar_parameters(mediaan_scenario['results'], participant_parameters)
        
        median_results = []
        for scenario in range(1, scenario_aantal + 1):
            try:
                median_results.append(process_scenario(
                    eerste_berekeningsjaar,
                    scenario, 
                    nummer, 
                    pensioenjaar_parameters, 
                    fourds, 
                    twods, 
                    endprognoseyear, 
                    pensioenleeftijd, 
                    pensioenjaar, 
                    pensioenjaar_parameters['calculationdate'].year
                ))
            except Exception as e:
                logging.error(f"Error processing median scenario {scenario} for participant {nummer}: {e}")
                raise e
        
        restructured_results = defaultdict(lambda: defaultdict(dict))

        for result in median_results:
            scenario = result["scenario"]
            for entry in result["scenarioresults"]:
                year = entry["jaar"]
                nominal_benefit = entry["nominal_benefit"]
                nominal_benefit_sr = entry["nominal_benefit_sr"]
                real_benefit = entry["real_benefit"]
                total_capital = entry["total_capital"]
                savings_op = entry["savings_op"]
                savings_hon = entry["savings_hon"]
                leeftijd = entry["leeftijd"]
                restructured_results[year][scenario] = {
                    "nominal_benefit": nominal_benefit,
                    "nominal_benefit_sr": nominal_benefit_sr,
                    "real_benefit": real_benefit,
                    "total_capital": total_capital,
                    "savings_op": savings_op,
                    "savings_hon": savings_hon,
                    "leeftijd": leeftijd
                    }
        
        final_result = {}
        percentiles = [5, 50, 95]

        for year, year_data in restructured_results.items():
            final_result[year] = get_percentiles_now(year_data, percentiles)

        
        return {
            "deelnemer": nummer + 1,
            "results": final_result
        }
    except Exception as e:
        logging.error(f"Unexpected error for participant {nummer}: {e}")
        return None
   
    
def read_input(self):
    return (parse_fourds(self), vertical_parameters(self), read_and_parse_twod(self))

def get_first_year(fourd_filtered, twod_filtered):
    fourd_years = [int(entry['year']) for entry in fourd_filtered if 'year' in entry]
    twod_years = [int(entry['year']) for entry in twod_filtered if 'year' in entry]
    if fourd_years and twod_years:
        prognosejaar = min(min(fourd_years), min(twod_years))
    elif fourd_years:
        prognosejaar = min(fourd_years)
    elif twod_years:
        prognosejaar = min(twod_years)
    else:
        prognosejaar = 1  # Default value if no year found
        
    return prognosejaar

def max_scenario(self):
    scenarios_fourd = [entry['scenario'] for entry in self.fourd]
    scenarios_twod = [entry['scenario'] for entry in self.twod]
    return max(scenarios_fourd + scenarios_twod)

def process_scenario(eerste_berekeningsjaar, scenario, nummer, parameters, fourd, twod, endprognoseyear, pensioenleeftijd, pensioenjaar, berekeningsjaar):
    prognosejaar = get_first_year(fourd, twod)
    return ontwikkel_scenario(
        eerste_berekeningsjaar=eerste_berekeningsjaar,
        endprognoseyear=endprognoseyear,
        scenario=scenario,
        parameters=parameters,
        fourd=fourd,
        twod=twod,
        pensioenleeftijd=pensioenleeftijd,
        prognosejaar=prognosejaar,
        pensioenjaar=pensioenjaar,
        eerste_jaar=berekeningsjaar,
        berekeningsjaar=berekeningsjaar
    )

def filter_scenario_data(fourd, twod, scenario):
    fourd_filtered = [entry for entry in fourd if entry["scenario"] == scenario]
    twod_filtered = [entry for entry in twod if entry["scenario"] == scenario]
    return fourd_filtered, twod_filtered

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.endprognoseyear = 2073
        self.filename = choose_file(self, 'input bestand')
        self.twod_sheet_naam = '2d'
        self.parameters_sheet_naam = 'Parameters'
        self.setGeometry(0, 0, 1800, 1100)
        # Add the QLabel for the loading message
        self.sheet_error_label = QLabel(self)
        self.sheet_error_label.setText("Input bestand heeft onjuiste format")
        self.sheet_error_label.setGeometry(500, 100, 1000, 400)
        self.choose_font(self.sheet_error_label, 30)
        self.sheet_error_label.hide()
        self.sheet_names_clicked = False
        self.fillDialog()
    
    def choose_font(self, item, f):
        font = QFont()
        font.setPointSize(f)
        item.setFont(font)
    def fillDialog(self):
        
        y2 = 50
        dy = 40
        
        y2 += dy
        self.sheets_button = QPushButton(self)
        self.sheets_button.setText("Maak URM prognose (stap 1)")
        self.sheets_button.setGeometry(5, y2, 250, 30)
        self.sheets_button.clicked.connect(self.readSheets)
        y2 += dy
        self.sheets_button2 = QPushButton(self)
        self.sheets_button2.setText("Maak URM prognose (stap 1)")
        self.sheets_button2.setGeometry(5, y2, 350, 30)

        self.sheets_button2.hide()
        self.end_button = QPushButton(self)
        self.end_button.setText("End")
        self.end_button.clicked.connect(self.close)
        self.end_button.move(5, 5)
        
        # Add the QLabel for the loading message
        self.loading_label = QLabel(self)
        self.loading_label.setText("Momentje, gegevens aan het inlezen")
        self.loading_label.setGeometry(5, y2 + 40, 400, 30)
        self.choose_font(self.loading_label, 8)
        self.loading_label.hide()
        # Add the QLabel for the loading message
        self.loading_label2 = QLabel(self)
        self.loading_label2.setText("Ingelezen, nu aan het rekenen")
        self.loading_label2.setGeometry(5, y2 + 40, 400, 30)
        self.choose_font(self.loading_label2, 8)
        self.loading_label2.hide()
        # Add the QLabel for the loading message
        self.loading_label3 = QLabel(self)
        self.loading_label3.setText("Alles berekend en weggeschreven")
        self.loading_label3.setGeometry(5, y2 + 40, 400, 30)
        self.choose_font(self.loading_label3, 8)
        self.loading_label3.hide()
        y2 += dy
        
        self.resultaten_button = QLabel(self)
        self.resultaten_button.setGeometry(5, y2, 250, 30)  # Position and size of the button
        self.resultaten_button.setText("Resultaten weggeschreven")
        #self.resultaten_button.clicked.connect(self.follow_up)
        self.resultaten_button.hide()
        
        layout = QVBoxLayout()
        
        layout.addWidget(self.end_button)
        layout.addWidget(self.loading_label)
        layout.addWidget(self.loading_label2)
        layout.addWidget(self.loading_label3)
        self.setLayout(layout)
        
    def readSheets(self):
        self.sheet_names_clicked = True
        self.loading_label.show()
        QApplication.processEvents() 
        self.twod_sheet = parse_sheet(self, self.filename, self.twod_sheet_naam, 1)
        if self.twod_sheet is None:
            self.sheet_error_label.show()
        self.onSheetsRead()
    def onSheetsRead(self):
        if self.sheet_names_clicked:
            self.parameters_variables = [
                "urm state", "savings start", "savings start honorary", "benefit start", "pensionbase",
                "birthdate", "startdate prognosis", "calculationdate", "enddate prognosis",
                "default retirement date", "age fraction", "year fraction retirement year",
                "amount of scenarios", "pp ratio", "start year", "person id"
            ]
            self.deelnemers = []
            time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"Input lezen om {time}")
            (fourd, parameters, twod) = read_input(self)
            self.fourd = fourd
            self.parameters = parameters
            self.eerste_berekeningsjaar = self.parameters[0]['calculationdate'].year
            self.twod = twod
            print(f"Input gelezen om {time}")
            self.loading_label.hide()
            self.loading_label2.show()
            self.calculate_stap1()
            self.loading_label2.hide()
            QApplication.processEvents()
    
    def calculate_stap1(self):
        self.deelnemerresults = []
        self.deelnemer_mediaan_results = []
        self.percentile_results = []
        self.pensioenleeftijd = 67
        
        self.scenario_aantal = max_scenario(self)
        self.deelnemers_aantal = len(self.parameters)
        self.contributionRate = 0.224
        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"start filtering: {time}")
        self.fourd_dict = {scenario: filter_scenario_data(self.fourd, self.twod, scenario)[0] for scenario in range(1, self.scenario_aantal)}
        self.twod_dict = {scenario: filter_scenario_data(self.fourd, self.twod, scenario)[1] for scenario in range(1, self.scenario_aantal)}
        print(f"aantal deelnemers: {self.deelnemers_aantal}")
        print(f"aantal scenarios: {self.scenario_aantal}")
        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"start porcessing: {time}")
        results = []
        for nummer in range(0,self.deelnemers_aantal):
           results.append(stap_1(nummer, self.eerste_berekeningsjaar, self.scenario_aantal, self.parameters, self.fourd, self.twod, self.endprognoseyear, self.pensioenleeftijd))
        
        results = [result for result in results if result is not None]
        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"End processing: {time}")
        self.deelnemerresults = results
        
        self.write_step_1()
        self.sheets_button2.setText("Bereken URM bedragen 10 jaar na pensioendatum (stap 2)")
        self.calculate_stap2()
        
    def calculate_stap2(self):
        print(f"aantal deelnemers: {self.deelnemers_aantal}")
        print(f"aantal scenarios: {self.scenario_aantal}")
        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"start time: {time}")
        results = []
        for idx, deelnemerresult in enumerate(self.deelnemerresults):
            results.append(stap_2(idx, self.eerste_berekeningsjaar, self.scenario_aantal, self.parameters, self.fourd, self.twod, self.endprognoseyear, self.pensioenleeftijd, self.deelnemerresults))
      
        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"end time: {time}")
        
        results = [result for result in results if result is not None]
        time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"Processed finished time: {time}")
        self.deelnemer_mediaan_results = results
        
        self.write_step_2()
        self.resultaten_button.show()
    
    def write_step_1(self):
        self.result_file_name = "results.xlsx"
        self.wb = openpyxl.Workbook()
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        ws = self.wb.create_sheet(title="Pensioendatum")
        ws["A1"] = "Deelnemer"
        ws["B1"] = "Pensioenjaar"
        ws["C1"] = "Savings 5pct"
        ws["D1"] = "Savings 50pct"
        ws["E1"] = "Savings 95pct"
        ws["F1"] = "Savings Honorary 5pct"
        ws["G1"] = "Savings Honorary 50pct"
        ws["H1"] = "Savings Honorary 95pct"
        ws["I1"] = "Benefits 5pct"
        ws["J1"] = "Benefits 50pct"
        ws["K1"] = "Benefits 95pct"
        ws["L1"] = "Benefits Deflated 5pct"
        ws["M1"] = "Benefits Deflated 50pct"
        ws["N1"] = "Benefits Deflated 95pct"
        ws["O1"] = "Scenario 5pct"
        ws["P1"] = "Scenario 50pct"
        ws["Q1"] = "Scenario 95pct"
        
        for deelnemer_nummer, result in enumerate(self.deelnemerresults):
            pensioenjaar = pensioenjaar_calculator(self.parameters[deelnemer_nummer], self.pensioenleeftijd)
            
            result = result['results']
            if pensioenjaar < self.parameters[deelnemer_nummer]['startdate prognosis'].year:
                pensioenjaar = self.parameters[deelnemer_nummer]['startdate prognosis'].year
            pensioenjaar = min(get_largest_year(result), pensioenjaar)
             
            rownum = deelnemer_nummer + 2
            ws[f"A{rownum}"] = deelnemer_nummer + 1
            ws[f"B{rownum}"] = pensioenjaar
            ws[f"C{rownum}"] = result[pensioenjaar]['5th']['results']['savings_op']
            ws[f"D{rownum}"] = result[pensioenjaar]['50th']['results']['savings_op']
            ws[f"E{rownum}"] = result[pensioenjaar]['95th']['results']['savings_op']
            ws[f"F{rownum}"] = result[pensioenjaar]['5th']['results']['savings_hon']
            ws[f"G{rownum}"] = result[pensioenjaar]['50th']['results']['savings_hon']
            ws[f"H{rownum}"] = result[pensioenjaar]['95th']['results']['savings_hon']
            ws[f"I{rownum}"] = result[pensioenjaar]['5th']['results']['nominal_benefit']
            ws[f"J{rownum}"] = result[pensioenjaar]['5th']['results']['nominal_benefit']
            ws[f"K{rownum}"] = result[pensioenjaar]['5th']['results']['nominal_benefit']
            ws[f"L{rownum}"] = result[pensioenjaar]['5th']['results']['real_benefit']
            ws[f"M{rownum}"] = result[pensioenjaar]['5th']['results']['real_benefit']
            ws[f"N{rownum}"] = result[pensioenjaar]['5th']['results']['real_benefit']
            ws[f"O{rownum}"] = result[pensioenjaar]['5th']['scenario']
            ws[f"P{rownum}"] = result[pensioenjaar]['50th']['scenario']
            ws[f"Q{rownum}"] = result[pensioenjaar]['95th']['scenario']
            
        print("Gegevens weggeschreven in results.xlsx")
        self.wb.save(self.result_file_name) 
    
    def write_step_2(self):
        ws = self.wb.create_sheet(title="10 jaar later")
        ws["A1"] = "Deelnemer"
        ws["B1"] = "Jaar"
        ws["C1"] = "Benefits 5pct"
        ws["D1"] = "Benefits 50pct"
        ws["E1"] = "Benefits 95pct"
        ws["F1"] = "Benefits Deflated 5pct"
        ws["G1"] = "Benefits Deflated 50pct"
        ws["H1"] = "Benefits Deflated 95pct"
        ws["I1"] = "Scenario 5pct"
        ws["J1"] = "Scenario 50pct"
        ws["K1"] = "Scenario 95pct"
        
        for deelnemer_nummer, result in enumerate(self.deelnemer_mediaan_results):
            pensioenjaar = min(self.endprognoseyear,10 + pensioenjaar_calculator(self.parameters[deelnemer_nummer], self.pensioenleeftijd))
            if pensioenjaar < self.parameters[deelnemer_nummer]['startdate prognosis'].year:
                pensioenjaar = self.parameters[deelnemer_nummer]['startdate prognosis'].year + 10
            largest_year = self.endprognoseyear -1
            #result_entry['results'][0]['scenarioresults'][self.parameters[deelnemer_nummer]['amount of scenarios']]['jaar']
            #get_largest_year(
            pensioenjaar = min(largest_year, pensioenjaar)
            print(f"Deelnemer {deelnemer_nummer + 1} heeft pensioenjaar + 10 {pensioenjaar}")
            results = result['results']
            rownum = deelnemer_nummer + 2
            ws[f"A{rownum}"] = deelnemer_nummer + 1
            ws[f"B{rownum}"] = pensioenjaar
            ws[f"C{rownum}"] = results[pensioenjaar]['5th']['results']['nominal_benefit']
            ws[f"D{rownum}"] = results[pensioenjaar]['50th']['results']['nominal_benefit']
            ws[f"E{rownum}"] = results[pensioenjaar]['95th']['results']['nominal_benefit']
            ws[f"F{rownum}"] = results[pensioenjaar]['5th']['results']['real_benefit']
            ws[f"G{rownum}"] = results[pensioenjaar]['50th']['results']['real_benefit']
            ws[f"H{rownum}"] = results[pensioenjaar]['95th']['results']['real_benefit']
            ws[f"I{rownum}"] = results[pensioenjaar]['5th']['scenario']
            ws[f"J{rownum}"] = results[pensioenjaar]['50th']['scenario']
            ws[f"K{rownum}"] = results[pensioenjaar]['95th']['scenario']
            
        print("Gegevens weggeschreven in results.xlsx")
        self.wb.save(self.result_file_name) 
        
   
if __name__ == "__main__":
    app = QApplication([])
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())