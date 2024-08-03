# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 12:30:27 2024

@author: M67B363
"""

from datetime import date
import openpyxl
import xlrd
import pandas as pd

def parse_sheet(self, filename, sheetname, skiprows):
    if filename:
        try:
            sheet = pd.read_excel(filename, sheet_name=sheetname, skiprows=skiprows)
            #workbook = openpyxl.load_workbook(filename, data_only=True, keep_vba=True)
            #sheet = workbook[sheetname]
            return sheet
        except KeyError:
            # Return None if the sheetname is not found
            return None
    return None

def parse_parameters(self, sheet, parameters_variables, column):
    start_row = 2
    result = {}
    for idx, param in enumerate(parameters_variables):
        
        cell = f'{column}{start_row + idx}'
        if param in ['birthdate', 'calculationdate', 'defalt retirement date', 'enddate prognosis', "default retirement date"]:
            cell_value = sheet[cell].value
            if isinstance(cell_value, int):
                date_value = xlrd.xldate_as_datetime(cell_value, 0)
                result[param] = date_value.date()
            else: 
                 result[param] = sheet[cell].value 
        else: result[param] = sheet[cell].value

    return result



def parse_twod(self, sheet):
    start_row = 2  # Assuming data starts from the second row
    result = {}

    row = start_row
    results = []
    
    while True:
        year = sheet[f'A{row}'].value
        if year is None or year == "" or not isinstance(year, (int, float)):
            break
        
        result = {
            "year": year,
            "scenario": sheet[f'B{row}'].value,
            "one_year_inflation": convert_to_float(sheet[f'C{row}'].value),
            "cpi": convert_to_float(sheet[f'D{row}'].value),
            "ff": convert_to_float(sheet[f'E{row}'].value),
            "payout_adjustment": convert_to_float(sheet[f'F{row}'].value),
            "sr_adjustment": convert_to_float(sheet[f'G{row}'].value),
            "contribution_rate": convert_to_float(sheet[f'H{row}'].value)
        }
        results.append(result)
        row += 1
 
    return results

def convert_to_float(value):
    if value is None or value == '' or str(value).isspace():
        return 0.0
    try:
        return float(value)
    except ValueError:
        return 0.0

def parse_fourds(self):
    sheetnames = ['18-27', '28-37', '38-47', '48-57', '58-67', '68-77', '78-87', '88-97', '98-107', '108-117', '118-127']
    fourds = []
    for sheetname in sheetnames:
        sheet = parse_sheet(self,self.filename, sheetname)
        four_d = parse_fourd(self, sheet)
        fourds.append(four_d)
        print(f"cohort {sheetname} parsed")
      
    return fourds

def parse_fourd(self, sheet):
    start_row = 2  # Assuming data starts from the second row
    result = {}

    row = start_row
    results = []
    
    while True:
        year = sheet[f'A{row}'].value
        if year is None or year == "" or not isinstance(year, (int, float)):
            break
        
        result = {
            "year": year,
            "scenario": sheet[f'B{row}'].value,
            "cohort": sheet[f'C{row}'].value,
            "cwf_op": convert_to_float(sheet[f'D{row}'].value),
            "cwf_pp": convert_to_float(sheet[f'E{row}'].value),
            "total_return": convert_to_float(sheet[f'F{row}'].value),
            "total_return_hon": convert_to_float(sheet[f'G{row}'].value)
        }
        results.append(result)
        row += 1
        
    return results

def parse_festina_resultaten(self, filename, status, accrual_row_count, benefit_row_count):
    accrual_sheet_name = f"{status}-accrualPhase"
    benefit_sheet_name = f"{status}-benefitPhase"
        
    accrual_sheet = parse_sheet(self, filename, accrual_sheet_name)
    if accrual_sheet is None:
        self.sheet_error_label.show()
    benefit_sheet = parse_sheet(self, filename, benefit_sheet_name)
    if benefit_sheet is None:
        self.sheet_error_label.show()
    accrual_rows = []
    benefit_rows = []
    if accrual_sheet is not None:
        sheet = accrual_sheet
        
        for rownumber in range(2, accrual_row_count + 1):
            row = {
                "leeftijd": sheet[f'A{rownumber}'].value,
                "prognosejaar": sheet[f'B{rownumber}'].value,
                "jaar": sheet[f'C{rownumber}'].value,
                "savings5": sheet[f'D{rownumber}'].value,
                "savings50": sheet[f'E{rownumber}'].value,
                "savings95": sheet[f'F{rownumber}'].value,
                "savings5_hon": sheet[f'G{rownumber}'].value,
                "savings50_hon": sheet[f'H{rownumber}'].value,
                "savings95_hon": sheet[f'I{rownumber}'].value,
                "scenario5": sheet[f'P{rownumber}'].value,
                "scenario50": sheet[f'Q{rownumber}'].value,
                "scenario95": sheet[f'R{rownumber}'].value,
                }
            accrual_rows.append(row)
    if benefit_sheet is not None:
        sheet = benefit_sheet
        
        for rownumber in range(2, benefit_row_count + 1):
            row = {
                "leeftijd": sheet[f'A{rownumber}'].value,
                "prognosejaar": sheet[f'B{rownumber}'].value,
                "jaar": sheet[f'C{rownumber}'].value,
                "benefit5": sheet[f'D{rownumber}'].value,
                "benefit50": sheet[f'E{rownumber}'].value,
                "benefit95": sheet[f'F{rownumber}'].value,
                "benefit5_def": sheet[f'G{rownumber}'].value,
                "benefit50_def": sheet[f'H{rownumber}'].value,
                "benefit95_def": sheet[f'I{rownumber}'].value,
                "scenario5": sheet[f'J{rownumber}'].value,
                "scenario50": sheet[f'K{rownumber}'].value,
                "scenario95": sheet[f'L{rownumber}'].value,
                }
            benefit_rows.append(row)
        
    return (accrual_rows, benefit_rows)